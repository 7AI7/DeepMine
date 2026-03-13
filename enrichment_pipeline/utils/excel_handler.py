# Enrichment Pipeline — Excel Input/Output
#
# Reads company names from input Excel.
# Writes enriched data with multi-row support (one company → multiple Maps locations).

import json
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import INPUT_EXCEL, INPUT_SHEET, OUTPUT_EXCEL, OUTPUT_DIR, PROGRESS_FILE


# ── Output column order ───────────────────────────────────────
OUTPUT_COLUMNS = [
    "company_name",
    "revenue_2023",
    "net_profit_2023",
    "employee_count",
    "tcc_source_url",
    "linkedin_url",
    "linkedin_followers",
    "maps_name",
    "address",
    "maps_link",
    "phone",
    "rating",
    "review_count",
    "email",
    "category",
]


def read_companies(filepath: str = INPUT_EXCEL, sheet: str = INPUT_SHEET) -> list[dict]:
    """
    Read company names from input Excel.
    
    Expected columns: 'name' (required), 'id' (optional — used for tracking).
    Returns list of {"name": "Company Name", "id": 123} dicts.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[sheet]
    
    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [h.lower().strip() if h else "" for h in headers]
    
    # Find column indices
    name_idx = None
    id_idx = None
    for i, h in enumerate(headers_lower):
        if h == "name":
            name_idx = i
        elif h == "id":
            id_idx = i
    
    if name_idx is None:
        raise ValueError(f"Column 'name' not found in sheet '{sheet}'. Headers: {headers}")
    
    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        if not name or not str(name).strip():
            continue
        company = {"name": str(name).strip()}
        if id_idx is not None and row[id_idx]:
            company["id"] = row[id_idx]
        companies.append(company)
    
    wb.close()
    return companies


def save_results(results: list[dict], filepath: str = OUTPUT_EXCEL) -> None:
    """
    Save enriched results to Excel.
    Creates the file with headers if it doesn't exist, appends otherwise.
    
    Each result dict should have keys matching OUTPUT_COLUMNS.
    Missing keys default to 'N/A'.
    """
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    
    p = Path(filepath)
    if p.exists():
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "enriched"
        # Write header row
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
    
    # Append data rows
    for result in results:
        row_data = []
        for col in OUTPUT_COLUMNS:
            row_data.append(result.get(col, "N/A"))
        ws.append(row_data)
    
    wb.save(filepath)
    wb.close()


def save_incremental(batch: list[dict], filepath: str = OUTPUT_EXCEL) -> None:
    """
    Append a batch of results to the output Excel.
    Wrapper around save_results for incremental saving.
    """
    save_results(batch, filepath)


# ── Progress tracking ─────────────────────────────────────────

def load_progress(filepath: str = PROGRESS_FILE) -> set[str]:
    """
    Load set of already-completed company names from progress file.
    Returns empty set if file doesn't exist.
    """
    p = Path(filepath)
    if not p.exists():
        return set()
    data = json.loads(p.read_text(encoding="utf-8"))
    return set(data.get("completed", []))


def save_progress(completed: set[str], filepath: str = PROGRESS_FILE) -> None:
    """
    Save progress — list of completed company names + timestamp.
    """
    data = {
        "completed": sorted(completed),
        "count": len(completed),
        "timestamp": datetime.now().isoformat(),
    }
    Path(filepath).write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )


def mark_completed(company_name: str, filepath: str = PROGRESS_FILE) -> None:
    """
    Add a single company name to the completed set and save.
    """
    completed = load_progress(filepath)
    completed.add(company_name)
    save_progress(completed, filepath)
