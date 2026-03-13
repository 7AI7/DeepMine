"""
Consolidate all result Excel files into a single final.xlsx.

Logic:
1. Keep existing final.xlsx data for IDs 100001-100400 (except 100020)
2. Replace 100020 with failed file data
3. Insert 62 missing failed IDs into 100001-100400 range
4. For IDs <= 203198: prefer failed file data over numbered file data
5. IDs > 203198: numbered files only
6. Re-generate sequential sub-sheet IDs (PR, A, C, M, I, MA)
"""
import openpyxl
from openpyxl import Workbook
import os
import shutil
import time
from collections import defaultdict

start_time = time.time()

BASE_DIR = r'D:\JAI\SHIVA FINAL - Copy\deep_crawler'
RESULT_DIR = os.path.join(BASE_DIR, 'result xls')
FINAL_PATH = os.path.join(BASE_DIR, 'final.xlsx')
BACKUP_PATH = os.path.join(BASE_DIR, 'final_backup.xlsx')
OUTPUT_PATH = os.path.join(BASE_DIR, 'final_new.xlsx')

# ============ TARGET COLUMN DEFINITIONS ============
MAIN_COLS = [
    'id','name','website','email','phone','address','city','states','country',
    'website_last_updated_on_year','infrastructure_available',
    'contact_person_designation','contact_person_name','contact_person_contact',
    'brochure_link','linkedin_page'
]
PRODUCTS_COLS = ['id','company_id','product_category','product','application','service','serving_sector']
ADDRESSES_COLS = ['id','company_id','address','city','states','country','address_label','pincode']
CLIENTS_COLS = ['id','company_id','client_name']
MANAGEMENT_COLS = ['id','company_id','designation','name','contact']
INFRA_COLS = ['id','company_id','block_name','capacity','equipment']
MACHINES_COLS = ['id','company_id','machine_name','brand_name','qty','capacity_value','capacity_unit','extra','specification']

# ============ HELPER FUNCTIONS ============
def read_sheet_rows(ws, forced_headers=None):
    """Generator: yields dicts from worksheet rows.
    If forced_headers provided, treats ALL rows as data (no header in file).
    """
    headers = forced_headers
    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        if headers is None:
            headers = list(row_vals)
            continue
        row_dict = {}
        for j, val in enumerate(row_vals):
            if j < len(headers) and headers[j] is not None:
                row_dict[headers[j]] = val
        yield row_dict

def map_row(row_dict, target_cols, aliases=None):
    """Map source row dict to target columns using aliases for name differences."""
    if aliases is None:
        aliases = {}
    result = {}
    for col in target_cols:
        if col in aliases:
            for alias in aliases[col]:
                if alias in row_dict:
                    result[col] = row_dict[alias]
                    break
            else:
                result[col] = row_dict.get(col)
        else:
            result[col] = row_dict.get(col)
    return result

# Column name aliases (target_name -> [possible_source_names])
MAIN_ALIASES = {'states': ['states', 'state']}
ADDR_ALIASES = {'states': ['states', 'state'], 'address_label': ['address_label', 'address_type']}

# ============ DATA STORAGE ============
all_companies = {}           # company_id -> {col: val}
all_products = defaultdict(list)
all_addresses = defaultdict(list)
all_clients = defaultdict(list)
all_management = defaultdict(list)
all_infra = defaultdict(list)
all_machines = defaultdict(list)
company_source = {}          # company_id -> source_name

# ============ PHASE 1: LOAD EXISTING FINAL.XLSX ============
print("Phase 1: Loading existing final.xlsx...")
wb = openpyxl.load_workbook(FINAL_PATH, read_only=True)

for row in read_sheet_rows(wb['main']):
    cid = row['id']
    all_companies[cid] = map_row(row, MAIN_COLS, MAIN_ALIASES)
    company_source[cid] = 'existing_final'

for row in read_sheet_rows(wb['Products']):
    all_products[row['company_id']].append(map_row(row, PRODUCTS_COLS))

for row in read_sheet_rows(wb['Addresses']):
    all_addresses[row['company_id']].append(map_row(row, ADDRESSES_COLS, ADDR_ALIASES))

for row in read_sheet_rows(wb['Clients']):
    all_clients[row['company_id']].append(map_row(row, CLIENTS_COLS))

for row in read_sheet_rows(wb['Management']):
    all_management[row['company_id']].append(map_row(row, MANAGEMENT_COLS))

for row in read_sheet_rows(wb['Infrastructure']):
    all_infra[row['company_id']].append(map_row(row, INFRA_COLS))

for row in read_sheet_rows(wb['Machines']):
    all_machines[row['company_id']].append(map_row(row, MACHINES_COLS))

wb.close()
existing_ids = set(all_companies.keys())
print(f"  Loaded {len(existing_ids)} companies (IDs {min(existing_ids)}-{max(existing_ids)})")
print(f"  Products: {sum(len(v) for v in all_products.values())}, "
      f"Addresses: {sum(len(v) for v in all_addresses.values())}, "
      f"Clients: {sum(len(v) for v in all_clients.values())}, "
      f"Management: {sum(len(v) for v in all_management.values())}, "
      f"Infrastructure: {sum(len(v) for v in all_infra.values())}, "
      f"Machines: {sum(len(v) for v in all_machines.values())}")

# ============ PHASE 2: LOAD FAILED FILE ============
print("\nPhase 2: Loading failed success file...")
wb = openpyxl.load_workbook(
    os.path.join(RESULT_DIR, '100001-203198 failed success.xlsx'), read_only=True)

# Load companies (only IDs <= 203198)
failed_company_data = {}
for row in read_sheet_rows(wb['companies']):
    cid = row['id']
    if cid <= 203198:
        failed_company_data[cid] = map_row(row, MAIN_COLS, MAIN_ALIASES)

failed_cids = set(failed_company_data.keys())
print(f"  Found {len(failed_cids)} failed companies (IDs <= 203198)")

# Load sub-sheets (only for relevant company IDs)
fail_sub = {
    'products': defaultdict(list),
    'addresses': defaultdict(list),
    'clients': defaultdict(list),
    'management': defaultdict(list),
    'infra': defaultdict(list),
    'machines': defaultdict(list),
}

sheet_configs = [
    ('products',  'products',              PRODUCTS_COLS,   None),
    ('addresses', 'addresses',             ADDRESSES_COLS,  ADDR_ALIASES),
    ('clients',   'clients',               CLIENTS_COLS,    None),
    ('management','management',            MANAGEMENT_COLS, None),
    ('infra',     'infrastructure_blocks', INFRA_COLS,      None),
    ('machines',  'machines',              MACHINES_COLS,   None),
]

for key, sheet_name, cols, aliases in sheet_configs:
    count = 0
    for row in read_sheet_rows(wb[sheet_name]):
        cid = row.get('company_id')
        if cid and cid in failed_cids:
            fail_sub[key][cid].append(map_row(row, cols, aliases))
            count += 1
    print(f"  Failed {key}: {count} rows")

wb.close()

# ============ PHASE 3: MERGE FAILED DATA ============
print("\nPhase 3: Merging failed data...")
replaced = 0
inserted = 0

for cid in sorted(failed_cids):
    if cid in existing_ids and cid != 100020:
        # Keep existing final data for all existing IDs except 100020
        continue

    # Use failed data
    all_companies[cid] = failed_company_data[cid]
    company_source[cid] = 'failed'
    all_products[cid] = fail_sub['products'].get(cid, [])
    all_addresses[cid] = fail_sub['addresses'].get(cid, [])
    all_clients[cid] = fail_sub['clients'].get(cid, [])
    all_management[cid] = fail_sub['management'].get(cid, [])
    all_infra[cid] = fail_sub['infra'].get(cid, [])
    all_machines[cid] = fail_sub['machines'].get(cid, [])

    if cid == 100020:
        replaced += 1
    else:
        inserted += 1

print(f"  Replaced: {replaced} (ID 100020)")
print(f"  Inserted: {inserted} new failed IDs")
print(f"  Total companies: {len(all_companies)}")

# Free memory
del failed_company_data, fail_sub

# ============ PHASE 4: LOAD NUMBERED FILES ============
numbered_files = [
    '100401-200249.xlsx',
    '200250-200650.xlsx',
    '200653-201398.xlsx',
    '201412-201909.xlsx',
    '201911-203198.xlsx',
    '203200-204636.xlsx',
    '204637-207215.xlsx',
]

for fname in numbered_files:
    fpath = os.path.join(RESULT_DIR, fname)
    print(f"\nPhase 4: Loading {fname}...")
    wb = openpyxl.load_workbook(fpath, read_only=True)

    # Load companies
    added_cids = set()
    skipped = 0
    for row in read_sheet_rows(wb['companies']):
        cid = row['id']
        if cid in all_companies:
            skipped += 1
            continue
        all_companies[cid] = map_row(row, MAIN_COLS, MAIN_ALIASES)
        company_source[cid] = fname
        added_cids.add(cid)

    print(f"  Companies: added={len(added_cids)}, skipped={skipped}")

    if not added_cids:
        wb.close()
        continue

    # Load sub-sheets (only for companies sourced from THIS file)
    sub_configs = [
        ('products',  'products',              all_products,    PRODUCTS_COLS,   None,         None),
        ('addresses', 'addresses',             all_addresses,   ADDRESSES_COLS,  ADDR_ALIASES, None),
        ('clients',   'clients',               all_clients,     CLIENTS_COLS,    None,         None),
        ('management','management',            all_management,  MANAGEMENT_COLS, None,         None),
        ('infra',     'infrastructure_blocks', all_infra,       INFRA_COLS,      None,         None),
        ('machines',  'machines',              all_machines,    MACHINES_COLS,   None,         None),
    ]

    for key, sheet_name, data_dict, cols, aliases, forced_hdr in sub_configs:
        # Special case: 200250-200650.xlsx addresses has no header row
        if fname == '200250-200650.xlsx' and sheet_name == 'addresses':
            forced_hdr = ['id', 'company_id', 'address_type', 'address', 'city', 'state', 'country', 'pincode']

        count = 0
        for row in read_sheet_rows(wb[sheet_name], forced_headers=forced_hdr):
            cid = row.get('company_id')
            if cid and cid in added_cids:
                data_dict[cid].append(map_row(row, cols, aliases))
                count += 1
        print(f"  {key}: {count} rows")

    wb.close()
    print(f"  Total companies so far: {len(all_companies)}")

# ============ PHASE 5: WRITE OUTPUT ============
print("\n" + "="*60)
print("Phase 5: Writing output...")
sorted_ids = sorted(all_companies.keys())
total_companies = len(sorted_ids)
print(f"Total companies: {total_companies}")
print(f"ID range: {sorted_ids[0]} - {sorted_ids[-1]}")

# Count totals for each sub-sheet
totals = {
    'products': sum(len(all_products.get(cid, [])) for cid in sorted_ids),
    'addresses': sum(len(all_addresses.get(cid, [])) for cid in sorted_ids),
    'clients': sum(len(all_clients.get(cid, [])) for cid in sorted_ids),
    'management': sum(len(all_management.get(cid, [])) for cid in sorted_ids),
    'infra': sum(len(all_infra.get(cid, [])) for cid in sorted_ids),
    'machines': sum(len(all_machines.get(cid, [])) for cid in sorted_ids),
}
for k, v in totals.items():
    print(f"  {k}: {v} rows")

# Determine ID padding width
def id_width(count):
    return max(4, len(str(count)))

pr_w = id_width(totals['products'])
a_w = id_width(totals['addresses'])
c_w = id_width(totals['clients'])
m_w = id_width(totals['management'])
i_w = id_width(totals['infra'])
ma_w = id_width(totals['machines'])

# Create output workbook (write_only for performance)
wb_out = Workbook(write_only=True)

# --- Main sheet ---
print("  Writing main...")
ws_main = wb_out.create_sheet('main')
ws_main.append(MAIN_COLS)
for cid in sorted_ids:
    row = all_companies[cid]
    ws_main.append([row.get(col) for col in MAIN_COLS])
print(f"    {total_companies} rows written")

# --- Products ---
print("  Writing Products...")
ws = wb_out.create_sheet('Products')
ws.append(PRODUCTS_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_products.get(cid, []):
        counter += 1
        out = [f'PR{counter:0{pr_w}d}', cid]
        out.extend([row.get(col) for col in PRODUCTS_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# --- Addresses ---
print("  Writing Addresses...")
ws = wb_out.create_sheet('Addresses')
ws.append(ADDRESSES_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_addresses.get(cid, []):
        counter += 1
        out = [f'A{counter:0{a_w}d}', cid]
        out.extend([row.get(col) for col in ADDRESSES_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# --- Clients ---
print("  Writing Clients...")
ws = wb_out.create_sheet('Clients')
ws.append(CLIENTS_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_clients.get(cid, []):
        counter += 1
        out = [f'C{counter:0{c_w}d}', cid]
        out.extend([row.get(col) for col in CLIENTS_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# --- Management ---
print("  Writing Management...")
ws = wb_out.create_sheet('Management')
ws.append(MANAGEMENT_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_management.get(cid, []):
        counter += 1
        out = [f'M{counter:0{m_w}d}', cid]
        out.extend([row.get(col) for col in MANAGEMENT_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# --- Infrastructure ---
print("  Writing Infrastructure...")
ws = wb_out.create_sheet('Infrastructure')
ws.append(INFRA_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_infra.get(cid, []):
        counter += 1
        out = [f'I{counter:0{i_w}d}', cid]
        out.extend([row.get(col) for col in INFRA_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# --- Machines ---
print("  Writing Machines...")
ws = wb_out.create_sheet('Machines')
ws.append(MACHINES_COLS)
counter = 0
for cid in sorted_ids:
    for row in all_machines.get(cid, []):
        counter += 1
        out = [f'MA{counter:0{ma_w}d}', cid]
        out.extend([row.get(col) for col in MACHINES_COLS[2:]])
        ws.append(out)
print(f"    {counter} rows written")

# Save
print("\n  Saving to disk...")
wb_out.save(OUTPUT_PATH)

elapsed = time.time() - start_time
print(f"\nDone! Saved to: {OUTPUT_PATH}")
print(f"Elapsed: {elapsed:.1f}s")

# ============ PHASE 6: VALIDATION ============
print("\n" + "="*60)
print("VALIDATION SUMMARY")
print("="*60)

# Source breakdown
sources = defaultdict(int)
for cid in sorted_ids:
    sources[company_source[cid]] += 1
print("\nCompanies by source:")
for src, cnt in sorted(sources.items()):
    print(f"  {src}: {cnt}")

# Check ID 100020 source
print(f"\nID 100020 source: {company_source.get(100020, 'NOT FOUND')}")

# Check a few milestone IDs
milestones = [100001, 100020, 100400, 100401, 100506, 200001, 200250, 203198, 203200, 207215]
print("\nMilestone ID check:")
for mid in milestones:
    if mid in all_companies:
        c = all_companies[mid]
        print(f"  {mid}: name={c.get('name','')[:40]}, source={company_source.get(mid)}")
    else:
        print(f"  {mid}: NOT PRESENT")

# First and last IDs
print(f"\nFirst company ID: {sorted_ids[0]}")
print(f"Last company ID: {sorted_ids[-1]}")
print(f"Total unique company IDs: {len(sorted_ids)}")
