import openpyxl

# Check overlaps between failed file and numbered files
wb_fail = openpyxl.load_workbook('result xls/100001-203198 failed success.xlsx', read_only=True)
ws_fail = wb_fail['companies']
failed_ids = set()
for i, row in enumerate(ws_fail.iter_rows(min_col=1, max_col=1, values_only=True)):
    if i == 0: continue
    failed_ids.add(row[0])
wb_fail.close()

# IDs in numbered files (excluding 100001-100400 already done)
numbered_ids = set()
for fname in ['100401-200249.xlsx', '200250-200650.xlsx', '200653-201398.xlsx', 
              '201412-201909.xlsx', '201911-203198.xlsx']:
    wb = openpyxl.load_workbook(f'result xls/{fname}', read_only=True)
    ws = wb['companies']
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0: continue
        numbered_ids.add(row[0])
    wb.close()

overlap = failed_ids & numbered_ids
failed_lte_203198 = {x for x in failed_ids if x <= 203198}
failed_only = failed_lte_203198 - numbered_ids

# Exclude 100001-100400 range already in final
wb_first = openpyxl.load_workbook('result xls/100001-100400.xlsx', read_only=True)
first_ids = set()
for i, row in enumerate(wb_first['companies'].iter_rows(min_col=1, max_col=1, values_only=True)):
    if i == 0: continue
    first_ids.add(row[0])
wb_first.close()

failed_only_excl_done = failed_only - first_ids

print(f'Failed IDs <= 203198: {len(failed_lte_203198)}')
print(f'  Also in numbered files (100401-203198): {len(overlap)}')
print(f'  Sample overlap: {sorted(overlap)[:20]}')
print(f'  ONLY in failed (not in any numbered): {len(failed_only)}')
print(f'  ONLY in failed (excl 100001-100400 range): {len(failed_only_excl_done)}')

# Check how many failed IDs overlap with first file
failed_in_first_range = failed_ids & first_ids
print(f'\nFailed IDs also in 100001-100400.xlsx: {len(failed_in_first_range)}')
print(f'  These: {sorted(failed_in_first_range)}')

# Failed IDs in 100001-100506 not in first file
failed_100_range = {x for x in failed_ids if 100001 <= x <= 100506}
only_failed_100 = failed_100_range - first_ids
print(f'\nFailed IDs 100001-100506 NOT in 100001-100400.xlsx: {len(only_failed_100)}')
print(f'  These: {sorted(only_failed_100)}')
