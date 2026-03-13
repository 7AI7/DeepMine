import openpyxl

# 215 failed IDs overlap with numbered files 100401-203198
# I need to understand: are these SAME data or is failed data preferred?
# Let's check a sample overlap - ID 202406

wb_fail = openpyxl.load_workbook('result xls/100001-203198 failed success.xlsx', read_only=True)
ws_fail = wb_fail['companies']
fail_data = {}
for i, row in enumerate(ws_fail.iter_rows(values_only=True)):
    if i == 0:
        fail_headers = list(row)
        continue
    fail_data[row[0]] = list(row)

# Find where 202406 sits in numbered files
for fname in ['201911-203198.xlsx']:
    wb = openpyxl.load_workbook(f'result xls/{fname}', read_only=True)
    ws = wb['companies']
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            num_headers = list(row)
            continue
        if row[0] == 202406:
            print(f"=== 202406 in {fname} ===")
            for j, h in enumerate(num_headers):
                if h and j < len(row):
                    print(f"  {h}: {row[j]}")
            break
    wb.close()

if 202406 in fail_data:
    print(f"\n=== 202406 in failed file ===")
    for j, h in enumerate(fail_headers):
        if h and j < len(fail_data[202406]):
            print(f"  {h}: {fail_data[202406][j]}")

wb_fail.close()

# Now check: which range do these 215 overlaps fall into?
failed_ids_full = set(fail_data.keys())
numbered_ids = set()
for fname in ['100401-200249.xlsx', '200250-200650.xlsx', '200653-201398.xlsx', 
              '201412-201909.xlsx', '201911-203198.xlsx']:
    wb = openpyxl.load_workbook(f'result xls/{fname}', read_only=True)
    ws = wb['companies']
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0: continue
        numbered_ids.add(row[0])
    wb.close()

overlap = sorted(failed_ids_full & numbered_ids)
print(f"\n=== All 215 overlapping IDs range ===")
print(f"Min: {min(overlap)}, Max: {max(overlap)}")

# Which numbered files do they belong to?
for fname in ['100401-200249.xlsx', '200250-200650.xlsx', '200653-201398.xlsx', 
              '201412-201909.xlsx', '201911-203198.xlsx']:
    wb = openpyxl.load_workbook(f'result xls/{fname}', read_only=True)
    ws = wb['companies']
    fids = set()
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0: continue
        fids.add(row[0])
    wb.close()
    ov = fids & set(overlap)
    if ov:
        print(f"  {fname}: {len(ov)} overlaps")
