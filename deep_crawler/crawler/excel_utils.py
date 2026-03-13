"""
crawler/excel_utils.py

Excel-specific utilities for fetching and saving extraction results.
Handles URL normalization, ID generation, and bulk Excel operations.
"""

from __future__ import annotations
import pandas as pd
from pathlib import Path
from typing import List, Tuple, Dict, Any
from urllib.parse import urlparse
import logging
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

LOG = logging.getLogger(__name__)


def normalize_url_to_homepage(url: str) -> str:
    """
    Convert any URL to its homepage format.

    Examples:
        http://xyz.com/products/abc → http://xyz.com
        https://www.abc.in/contact.html → https://www.abc.in

    Args:
        url: Original URL (may have paths)

    Returns:
        Homepage URL (scheme + domain only)
    """
    url = url.strip()

    # Add scheme if missing
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url

    try:
        parsed = urlparse(url)
        # Return scheme + netloc only (no path)
        return f"{parsed.scheme}://{parsed.netloc}"
    except Exception as e:
        LOG.warning(f"Failed to normalize URL: {url}, error: {e}")
        return url


def fetch_sites_from_excel(
    excel_path: str,
    sheet_name: str,
    start_id: str,
    end_id: str
) -> List[Tuple[int, str, str]]:
    """
    Fetch company data from Excel sheet.
    
    Excel format:
    - Column A: Company ID (integer, e.g., 100018, 100019, ...)
    - Column B: Website URL (string)

    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet name (e.g., "Sheet1")
        start_id: Start ID like "S0001" or "100018"
        end_id: End ID like "S0050" or "100100"

    Returns:
        List of (company_id: int, normalized_url: str, original_url: str)
        company_id is READ from Excel Column A

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet doesn't exist or columns missing
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        # Only read Column B (website URLs), Column A is ignored
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[1])
    except Exception as e:
        raise ValueError(f"Failed to read sheet '{sheet_name}': {e}")

    # Validate column exists
    if len(df.columns) < 1:
        raise ValueError(f"Sheet '{sheet_name}' must have at least Column B (Website)")

    # Rename for easier access
    df.columns.values[0] = 'Website'
    
    # Parse row range from start_id/end_id
    # "S0001" → row 1, "S0050" → row 50
    def parse_row_number(id_str: str) -> int:
        id_str = str(id_str).strip().upper()
        if id_str.startswith('S'):
            # S0001 → extract 1
            return int(id_str[1:])
        else:
            # Plain numeric "100018" → treat as-is (legacy support)
            # Subtract 200000 to get row number if >= 200000
            numeric = int(id_str)
            if numeric >= 200000:
                return numeric - 200000
            else:
                return numeric

    start_row = parse_row_number(start_id)
    end_row = parse_row_number(end_id)

    # Generate row numbers (1-based, matching Excel row numbers after header)
    df['RowNumber'] = range(1, len(df) + 1)
    
    # Filter rows in range [start_row, end_row]
    df_slice = df[(df['RowNumber'] >= start_row) & (df['RowNumber'] <= end_row)].copy()

    if df_slice.empty:
        total_rows = len(df)
        raise ValueError(
            f"No companies found in rows {start_id} ({start_row}) "
            f"to {end_id} ({end_row}). "
            f"Total rows available in Excel: {total_rows}"
        )
    
    # Generate company IDs: 200000 + row_number
    df_slice['ID'] = 200000 + df_slice['RowNumber']

    # Extract company data
    results = []
    for idx, row in df_slice.iterrows():
        company_id = int(row["ID"])  # Already generated as 200000 + row_number
        row_number = int(row["RowNumber"])

        original_url = str(row["Website"]).strip()
        if not original_url or original_url.lower() == 'nan':
            LOG.warning(f"Skipping row {row_number} (ID {company_id}): empty website")
            continue

        # Handle Excel HYPERLINK formulas: =HYPERLINK("url", "text")
        if original_url.startswith('=HYPERLINK'):
            import re
            match = re.search(r'=HYPERLINK\("([^"]+)"', original_url)
            if match:
                original_url = match.group(1)
            else:
                LOG.warning(f"Row {row_number} (ID {company_id}): couldn't parse HYPERLINK formula")
                continue

        normalized_url = normalize_url_to_homepage(original_url)
        results.append((company_id, normalized_url, original_url))

    LOG.info(
        f"Fetched {len(results)} companies from Excel '{sheet_name}' "
        f"(rows {start_row}-{end_row}, IDs {200000+start_row}-{200000+end_row}): ")

    return results

def save_results_to_excel(
    excel_path: str,
    results: Dict[int, Dict[str, Any]]
) -> None:
    """
    Save extraction results to Excel output sheets.

    Args:
        excel_path: Path to Excel file
        results: Dict of {company_id: extraction_result_dict}

    Process:
        1. Clear all output sheets (companies, Products, Addresses, etc.)
        2. Parse results and generate child table IDs
        3. Write to respective sheets
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    LOG.info(f"Saving results to Excel: {excel_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path)

    # Sheet names (must match your Excel structure)
    sheet_names = [
        "companies",
        "Products",
        "Addresses",
        "Clients",
        "Management",
        "Infrastructure",
        "Machines"
    ]

    # Clear all output sheets (keep headers)
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Delete all rows except header (row 1)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

    # Initialize ID counters for child tables

    counters = {
        "products": 1,
        "addresses": 1,
        "clients": 1,
        "management": 1,
        "infrastructure": 1,
        "machines": 1
    }

    id_prefixes = {
        "products": "PR",
        "addresses": "A",
        "clients": "C",
        "management": "M",
        "infrastructure": "I",
        "machines": "MC"
    }


    # Prepare DataFrames for each sheet
    companies_rows = []
    products_rows = []
    addresses_rows = []
    clients_rows = []
    management_rows = []
    infrastructure_rows = []
    machines_rows = []

    for company_id, result in sorted(results.items()):
        try:
            
            # Companies table
            company_data = result.get("company", {}) or {}
            phone_raw = company_data.get("phone")
            if isinstance(phone_raw, dict):
                # Extract all phone values and join with comma+space
                phone_vals = [str(v) for v in phone_raw.values() if v]
                phone_str = ", ".join(phone_vals)
            elif isinstance(phone_raw, str):
                phone_str = phone_raw
            else:
                phone_str = None
            companies_rows.append({
                "id": company_id,
                "name": company_data.get("name"),
                "website": company_data.get("website"),
                "email": company_data.get("email"),
                "phone": phone_str,
                "address": company_data.get("address"),
                "city": company_data.get("city"),
                "state": company_data.get("state"),
                "country": company_data.get("country"),
                "google_maps_link": company_data.get("google_maps_link"),
                "google_address": company_data.get("google_address"),
                "Google_ratings": company_data.get("Google_ratings"),
                "Google_phoneNo.": company_data.get("Google_phoneNo."),
                "Google_email": company_data.get("Google_email"),
                "contact_name": company_data.get("contact_name"),
                "website_last_updated_on_year": company_data.get("website_last_updated_on_year"),
                "linkedin_page": company_data.get("linkedin_page"),
                "linkedin_employee_range": company_data.get("linkedin_employee_range"),
                "linkedin_followers_range": company_data.get("linkedin_followers_range"),
                "jd_contact_no": company_data.get("jd_contact_no"),
                "jd_contact_name": company_data.get("jd_contact_name"),
                "jd_email_id": company_data.get("jd_email_id"),
                "turnover_fy23": company_data.get("turnover_fy23"),
                "turnover_fy24": company_data.get("turnover_fy24"),
                "turnover_fy25": company_data.get("turnover_fy25"),
                "contact_person_designation": company_data.get("contact_person_designation"),
                "contact_person_name": company_data.get("contact_person_name"),
                "contact_person_contact": company_data.get("contact_person_contact"),
                "infrastructure_available": company_data.get("infrastructure_available"),
                "brochure_link": company_data.get("brochure_link")
            })

            # Products table
            products_data = result.get("products", {}) or {}

            if isinstance(products_data, list):
                # GLM Format: List of dicts with nested lists
                # Example: [{"product": ["a", "b"], "category": ["x", "y"]}]
                for product_obj in products_data:
                    if not product_obj or not isinstance(product_obj, dict):
                        continue

                    # Extract all fields from this product object
                    categories = product_obj.get("product_category") or product_obj.get("category") or []
                    products = product_obj.get("product") or product_obj.get("name") or []
                    applications = product_obj.get("application") or []
                    services = product_obj.get("service") or []
                    sectors = product_obj.get("serving_sector") or product_obj.get("sector") or []

                    # Ensure all are lists
                    if not isinstance(categories, list):
                        categories = [categories] if categories else []
                    if not isinstance(products, list):
                        products = [products] if products else []
                    if not isinstance(applications, list):
                        applications = [applications] if applications else []
                    if not isinstance(services, list):
                        services = [services] if services else []
                    if not isinstance(sectors, list):
                        sectors = [sectors] if sectors else []

                    # Find max length to iterate
                    max_len = max(
                        len(categories), len(products), len(applications),
                        len(services), len(sectors)
                    )

                    # Create rows by zipping all lists (pad shorter ones with None)
                    for i in range(max_len):
                        products_rows.append({
                            "id": f"{id_prefixes['products']}{counters['products']:04d}",
                            "company_id": company_id,
                            "product_category": categories[i] if i < len(categories) else None,
                            "product": products[i] if i < len(products) else None,
                            "application": applications[i] if i < len(applications) else None,
                            "service": services[i] if i < len(services) else None,
                            "serving_sector": sectors[i] if i < len(sectors) else None
                        })
                        counters["products"] += 1

            elif isinstance(products_data, dict):
                # Dict format: {"product_category": [...], "product": [...], ...}
                for key in ["product_category", "product", "application", "service", "serving_sector"]:
                    items = products_data.get(key, []) or []
                    if not isinstance(items, list):
                        items = [items] if items else []

                    for item in items:
                        if item:
                            products_rows.append({
                                "id": f"{id_prefixes['products']}{counters['products']:04d}",
                                "company_id": company_id,
                                "product_category": item if key == "product_category" else None,
                                "product": item if key == "product" else None,
                                "application": item if key == "application" else None,
                                "service": item if key == "service" else None,
                                "serving_sector": item if key == "serving_sector" else None
                            })
                            counters["products"] += 1

            # Addresses table
            addresses_data = result.get("addresses", []) or []
            if not isinstance(addresses_data, list):
                addresses_data = [addresses_data] if addresses_data else []

            for addr in addresses_data:
                if addr and isinstance(addr, dict):
                    addresses_rows.append({
                        "id": f"{id_prefixes['addresses']}{counters['addresses']:04d}",
                        "company_id": company_id,
                        "address": addr.get("address"),
                        "city": addr.get("city"),
                        "states": addr.get("state"),
                        "country": addr.get("country"),
                        "address_label": addr.get("label"),
                        "pincode": addr.get("pincode")
                    })
                    counters["addresses"] += 1

            # Clients table
            clients_data = result.get("clients", []) or []
            if not isinstance(clients_data, list):
                clients_data = [clients_data] if clients_data else []

            for client in clients_data:
                if client:
                    clients_rows.append({
                        "id": f"{id_prefixes['clients']}{counters['clients']:04d}",
                        "company_id": company_id,
                        "client_name": client if isinstance(client, str) else client.get("name")
                    })
                    counters["clients"] += 1

            # Management table
            management_data = result.get("management", []) or []
            if not isinstance(management_data, list):
                management_data = [management_data] if management_data else []

            for mgmt in management_data:
                if mgmt and isinstance(mgmt, dict):
                    management_rows.append({
                        "id": f"{id_prefixes['management']}{counters['management']:04d}",
                        "company_id": company_id,
                        "designation": mgmt.get("designation"),
                        "name": mgmt.get("name")
                    })
                    counters["management"] += 1

            # Infrastructure table
            infra_data = result.get("infrastructure")
            blocks = []

            if isinstance(infra_data, dict):
                # Old format: {"blocks": [{"name": "Block A", "machines": [...]}]}
                blocks = infra_data.get("blocks", []) or []
                if not isinstance(blocks, list):
                    blocks = [blocks] if blocks else []
                    
            elif isinstance(infra_data, list):
                # New format: LLM returned list of addresses with address_label
                # These are NOT infrastructure blocks, they're multiple company locations
                # Move them to Addresses table instead
                for addr in infra_data:
                    if isinstance(addr, dict) and addr.get("address"):
                        addresses_rows.append({
                            "id": f"{id_prefixes['addresses']}{counters['addresses']:04d}",
                            "company_id": company_id,
                            "address": addr.get("address"),
                            "city": addr.get("city"),
                            "states": addr.get("state"),
                            "country": addr.get("country"),
                            "address_label": addr.get("address_label"),  # NEW
                            "pincode": addr.get("pincode")
                        })
                        counters["addresses"] += 1
                blocks = []  # No infrastructure blocks in this case
            if not isinstance(blocks, list):
                blocks = [blocks] if blocks else []

            for block in blocks:
                if block:
                    block_name = block if isinstance(block, str) else block.get("name")
                    block_id = f"{id_prefixes['infrastructure']}{counters['infrastructure']:04d}",
                    infrastructure_rows.append({
                        "id": block_id,
                        "company_id": company_id,
                        "block_name": block_name
                    })
                    counters["infrastructure"] += 1

                    # Machines for this block
                    if isinstance(block, dict):
                        machines = block.get("machines", []) or []
                        if not isinstance(machines, list):
                            machines = [machines] if machines else []

                        for machine in machines:
                            if machine and isinstance(machine, dict):
                                machines_rows.append({
                                    "id": f"{id_prefixes['machines']}{counters['machines']:04d}",
                                    "company_id": company_id,
                                    "block_id": block_id,
                                    "machine_name": machine.get("machine_name"),
                                    "brand_name": machine.get("brand_name"),
                                    "qty": machine.get("qty"),
                                    "capacity_value": machine.get("capacity_value"),
                                    "capacity_unit": machine.get("capacity_unit"),
                                    "extra": machine.get("extra")
                                })
                                counters["machines"] += 1

        except Exception as e:
            LOG.exception(f"Failed to process company_id={company_id}: {e}")
            continue

    # Write DataFrames to sheets
    def write_df_to_sheet(sheet_name: str, rows: List[Dict]):
        if sheet_name not in wb.sheetnames:
            LOG.warning(f"Sheet '{sheet_name}' not found, skipping")
            return

        if not rows:
            LOG.info(f"No data for sheet '{sheet_name}'")
            return

        df = pd.DataFrame(rows)
        ws = wb[sheet_name]

        # Append rows 
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(list(row))

        LOG.info(f"Wrote {len(rows)} rows to sheet '{sheet_name}'")

    write_df_to_sheet("companies", companies_rows)
    write_df_to_sheet("Products", products_rows)
    write_df_to_sheet("Addresses", addresses_rows)
    write_df_to_sheet("Clients", clients_rows)
    write_df_to_sheet("Management", management_rows)
    write_df_to_sheet("Infrastructure", infrastructure_rows)
    write_df_to_sheet("Machines", machines_rows)

    # Save workbook
    wb.save(excel_path)
    LOG.info(f"Excel file saved successfully: {excel_path}")
