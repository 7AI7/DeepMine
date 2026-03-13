"""Excel file handling"""
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import config

class ExcelHandler:
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.workbook = None
        self.directors_buffer = []
        self.management_buffer = []
        
    def read_companies(self):
        """Read companies from input Excel"""
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")
        
        self.workbook = load_workbook(self.filepath)
        
        if config.INPUT_SHEET not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{config.INPUT_SHEET}' not found in Excel file")
        
        sheet = self.workbook[config.INPUT_SHEET]
        companies = []
        
        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # id and name
                companies.append({
                    "id": str(row[0]),
                    "name": str(row[1])
                })
        
        return companies
    
    def _ensure_directors_sheet(self):
        """Create directors sheet if it doesn't exist"""
        if self.workbook is None:
            self.workbook = load_workbook(self.filepath)
        
        if config.DIRECTORS_SHEET not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(config.DIRECTORS_SHEET)
            # Add headers
            for col_idx, header in enumerate(config.DIRECTORS_COLUMNS, start=1):
                sheet.cell(row=1, column=col_idx, value=header)
            self.workbook.save(self.filepath)
    
    def add_director_data(self, data):
        """Add data to directors buffer"""
        self.directors_buffer.append(data)
    
    def add_management_data(self, data):
        """Add data to management buffer"""
        self.management_buffer.append(data)
    
    def save_incremental(self):
        """Save buffered data to Excel"""
        if not self.directors_buffer and not self.management_buffer:
            return
        
        self._ensure_directors_sheet()
        self.workbook = load_workbook(self.filepath)
        
        # Save directors data
        if self.directors_buffer:
            directors_sheet = self.workbook[config.DIRECTORS_SHEET]
            for data in self.directors_buffer:
                row_data = [
                    data.get("company_id", ""),
                    data.get("person_name", ""),
                    data.get("designation", ""),
                    data.get("related_company", ""),
                    data.get("industry", ""),
                    data.get("status", ""),
                    data.get("Designation_in_other_company", ""),
                    data.get("contact", "")
                ]
                directors_sheet.append(row_data)
            self.directors_buffer.clear()
        
        # Save management data
        if self.management_buffer:
            if config.MANAGEMENT_SHEET in self.workbook.sheetnames:
                management_sheet = self.workbook[config.MANAGEMENT_SHEET]
                for data in self.management_buffer:
                    row_data = [
                        data.get("id", ""),
                        data.get("company_id", ""),
                        data.get("name", ""),
                        data.get("designation", ""),
                        data.get("contact", "")
                    ]
                    management_sheet.append(row_data)
            self.management_buffer.clear()
        
        self.workbook.save(self.filepath)
    
    def close(self):
        """Save remaining buffer and close"""
        self.save_incremental()
        if self.workbook:
            self.workbook.close()
